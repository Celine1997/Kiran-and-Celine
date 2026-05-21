from __future__ import annotations

import base64
import re
from dataclasses import dataclass
from datetime import datetime
from html import escape
from pathlib import Path

import openpyxl
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).parent
ASSETS = ROOT / "assets"
WORKBOOK_PATH = ROOT / "C&K.xlsx"
DEFAULT_SITE_PASSWORD = "K&C Wedding"
MUSIC_URL = "/app/static/Celine%20and%20Kiran%20V1.mp3"


@dataclass(frozen=True)
class Guest:
    name: str
    seats: int
    hindu: bool
    christian: bool


def image_uri(path: Path) -> str:
    if not path.exists():
        return ""
    encoded = base64.b64encode(path.read_bytes()).decode("utf-8")
    return f"data:image/png;base64,{encoded}"


def rows(sheet_name: str) -> list[list[object]]:
    workbook = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    sheet = workbook[sheet_name]
    return [
        list(row)
        for row in sheet.iter_rows(values_only=True)
        if any(cell is not None and str(cell).strip() for cell in row)
    ]


def first_column(sheet_name: str) -> list[str]:
    return [
        str(row[0]).strip()
        for row in rows(sheet_name)
        if row and row[0] is not None and str(row[0]).strip()
    ]


def parse_events() -> tuple[str, list[dict[str, str]]]:
    info = rows("Wedding Information")
    intro = str(info[0][0]).strip() if info else "We invite you to join us as we say I do."
    header_index = next((i for i, row in enumerate(info) if row and str(row[0]).strip() == "Events"), -1)
    if header_index < 0:
        return intro, []
    headers = [str(cell or "").strip() for cell in info[header_index]]
    events: list[dict[str, str]] = []
    for row in info[header_index + 1 :]:
        event = {}
        for index, header in enumerate(headers):
            value = row[index] if index < len(row) else ""
            event[header] = str(value).strip() if value is not None and str(value).strip() else "TBD"
        events.append(event)
    return intro, events


def parse_guests() -> list[Guest]:
    guests: list[Guest] = []
    for row in rows("Guest List")[1:]:
        name = str(row[0]).strip() if row and row[0] else ""
        if not name:
            continue
        seats = int(row[1] or 1) if len(row) > 1 else 1
        hindu = len(row) > 2 and str(row[2]).strip().lower() == "yes"
        christian = len(row) > 3 and str(row[3]).strip().lower() == "yes"
        guests.append(Guest(name, seats, hindu, christian))
    return guests


def normalize(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", value.lower()).strip()


def find_guest(query: str, guests: list[Guest]) -> Guest | None:
    needle = normalize(query)
    if not needle:
        return None
    for guest in guests:
        haystack = normalize(guest.name)
        if needle == haystack or needle in haystack:
            return guest
    return None


def timeline_items(text: str) -> list[tuple[str, str]]:
    compact = re.sub(r"\s+", " ", str(text or "")).strip()
    if not compact or compact == "TBD":
        return [("Time", "Details to be announced")]
    matches = re.findall(r"([^:]+):\s*([^:]+?)(?=\s+[A-Z][A-Za-z ]+:|$)", compact)
    if matches:
        return [(label.strip(), body.strip()) for label, body in matches]
    return [("Timeline", compact)]


def save_rsvp(
    guest: Guest,
    response: str,
    hindu_attending: bool,
    christian_attending: bool,
    contact_requested: bool,
) -> None:
    workbook = openpyxl.load_workbook(WORKBOOK_PATH)
    sheet_name = "RSVP Responses"
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(sheet_name)
    if sheet.max_row == 1 and not sheet.cell(1, 1).value:
        sheet.append([])
    if sheet.max_row < 1 or sheet.cell(1, 1).value != "Timestamp":
        sheet.delete_rows(1, sheet.max_row)
        sheet.append(
            [
                "Timestamp",
                "Guest Name",
                "Reserved Seats",
                "Response",
                "Hindu Wedding",
                "Christian Wedding",
                "Contact Directly",
            ]
        )

    sheet.append(
        [
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            guest.name,
            guest.seats,
            response,
            "Yes" if hindu_attending else "No",
            "Yes" if christian_attending else "No",
            "Yes" if contact_requested else "No",
        ]
    )

    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="7B2638")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    for column, width in {"A": 21, "B": 26, "C": 16, "D": 14, "E": 16, "F": 18, "G": 18}.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A2"
    workbook.save(WORKBOOK_PATH)


@st.cache_data(show_spinner=False)
def load_data() -> dict[str, object]:
    landing = first_column("Landing & Invitation")
    story = first_column("Our Love Story")
    registry = first_column("Registry")
    details_intro, events = parse_events()
    return {
        "couple": landing[1] if len(landing) > 1 else "Celine & Kiran",
        "invitation": landing[3] if len(landing) > 3 else "Join us as we begin our forever",
        "story": story,
        "registry": registry,
        "details_intro": details_intro,
        "events": events,
        "guests": parse_guests(),
        "logo_burgundy": image_uri(ASSETS / "logo-burgundy.png"),
        "logo_ivory": image_uri(ASSETS / "logo-ivory.png"),
        "hero": image_uri(ASSETS / "wedding-hero.png"),
    }


def css(data: dict[str, object]) -> str:
    return f"""
    <style>
    :root {{
      --ivory: #f8f5f2;
      --paper: #fffaf7;
      --linen: #ede6df;
      --champagne: #d6c3a3;
      --gold: #a97724;
      --charcoal: #2b2b2b;
      --muted: #6f625c;
      --burgundy: #7b2638;
      --burgundy-dark: #581725;
      --sage: #6f806c;
      --shadow: 0 28px 80px rgba(72, 41, 35, 0.14);
    }}

    html {{ scroll-behavior: smooth; scroll-padding-top: 92px; }}
    header[data-testid="stHeader"], div[data-testid="stToolbar"], div[data-testid="stDecoration"], footer {{ display: none; }}
    .block-container {{ max-width: none; padding: 0; }}
    .element-container {{ margin: 0 !important; }}
    .stApp {{ background: var(--ivory); color: var(--charcoal); font-family: Lato, Avenir, "Segoe UI", sans-serif; }}
    a, a:hover, a:visited, a:active {{ color: inherit; text-decoration: none !important; }}
    p {{ line-height: 1.65; }}

    .site-header {{
      position: fixed; top: 0; left: 0; right: 0; z-index: 30;
      min-height: 72px; padding: 10px clamp(18px, 4vw, 54px);
      display: flex; align-items: center; justify-content: space-between;
      color: var(--burgundy);
      border-bottom: 1px solid rgba(123, 38, 56, 0.1);
      background: linear-gradient(90deg, rgba(255,250,247,.98), rgba(248,245,242,.94));
      box-shadow: 0 12px 40px rgba(72, 41, 35, 0.08);
      backdrop-filter: blur(18px);
    }}
    .site-header::after {{ content: ""; position: absolute; left: 0; right: 0; bottom: -1px; height: 1px; background: linear-gradient(90deg, transparent, rgba(214,195,163,.72), transparent); }}
    .brand img {{ width: 58px; display: block; filter: drop-shadow(0 6px 14px rgba(88,23,37,.08)); }}
    .desktop-nav {{ display: flex; gap: clamp(18px, 3vw, 34px); font-size: .72rem; font-weight: 900; letter-spacing: .1em; text-transform: uppercase; }}
    .desktop-nav a {{ color: var(--burgundy) !important; border: 0 !important; transition: transform 180ms ease, color 180ms ease; }}
    .desktop-nav a:hover {{ color: var(--burgundy-dark) !important; transform: translateY(-1px); }}
    .nav-button, .sticky-rsvp, div.stButton > button {{
      display: inline-flex; align-items: center; justify-content: center;
      min-height: 44px; padding: 0 20px; border: 0; border-radius: 999px;
      background: var(--burgundy); color: #fff !important; font-weight: 900;
      box-shadow: 0 16px 38px rgba(88,23,37,.18);
      transition: transform 180ms ease, box-shadow 180ms ease, background 180ms ease;
    }}
    .nav-button:hover, .sticky-rsvp:hover, div.stButton > button:hover {{ background: var(--burgundy-dark); transform: translateY(-1px); }}

    .hero {{
      margin-top: 72px; min-height: calc(100vh - 72px);
      display: grid; align-items: center; overflow: hidden;
      padding: 74px clamp(22px, 5vw, 96px); color: #fff;
      background:
        radial-gradient(circle at 78% 24%, rgba(255,250,247,.30), transparent 31%),
        linear-gradient(90deg, rgba(55,17,26,.76), rgba(55,17,26,.28) 56%, rgba(55,17,26,.08)),
        linear-gradient(0deg, rgba(31,19,16,.42), rgba(31,19,16,.02)),
        url("{data["hero"]}") center / cover no-repeat;
    }}
    .hero-content {{ width: min(820px, 100%); }}
    .hero-logo {{ width: min(170px, 42vw); margin-bottom: 18px; opacity: .9; filter: drop-shadow(0 12px 26px rgba(37,18,17,.22)); animation: weddingFloat 7s ease-in-out infinite; }}
    .eyebrow, .event-label {{ margin: 0 0 14px; color: var(--gold); font-size: .75rem; font-weight: 900; letter-spacing: .15em; text-transform: uppercase; }}
    .hero .eyebrow {{ color: var(--champagne); }}
    h1, h2, h3 {{ margin: 0; font-family: Georgia, "Times New Roman", serif; font-weight: 700; letter-spacing: 0; line-height: .96; color: var(--burgundy); }}
    h1 {{ max-width: 8.4ch; color: #fff; font-size: clamp(4.4rem, 10.4vw, 8.8rem); line-height: .9; text-shadow: 0 14px 38px rgba(37,18,17,.24); }}
    h2 {{ font-size: clamp(2.15rem, 4.35vw, 3.8rem); }}
    h3 {{ font-size: clamp(1.8rem, 3.4vw, 2.8rem); }}
    .hero-subtitle {{ margin: 18px 0 28px; color: rgba(255,255,255,.88); font-size: clamp(1.12rem, 2vw, 1.45rem); font-weight: 700; }}
    .countdown {{ display: inline-grid; grid-template-columns: auto minmax(0,1fr); align-items: center; gap: 12px; min-height: 64px; padding: 10px 18px; border: 1px solid rgba(255,255,255,.3); border-radius: 999px; background: rgba(255,250,247,.16); backdrop-filter: blur(12px); }}
    .countdown-number {{ color: var(--champagne); font-family: Georgia, serif; font-size: clamp(2rem, 4vw, 3.2rem); font-weight: 700; line-height: .86; }}
    .countdown-label {{ color: rgba(255,255,255,.86); font-size: .76rem; font-weight: 900; letter-spacing: .08em; text-transform: uppercase; }}

    .section {{ scroll-margin-top: 92px; padding: clamp(76px, 9vw, 124px) clamp(18px, 6vw, 88px); }}
    .story {{ background: radial-gradient(circle at 10% 20%, rgba(234,214,207,.26), transparent 30%), linear-gradient(135deg, rgba(214,195,163,.22), rgba(248,245,242,0) 46%), var(--ivory); }}
    .gallery, .rsvp {{ background: #fff; }}
    .details {{ background: var(--linen); }}
    .registry {{ background: var(--ivory); }}
    .section-grid, .section-intro, .gallery-grid, .event-grid, .registry-layout {{ width: min(1120px, 100%); margin-inline: auto; }}
    .section-grid {{ display: grid; grid-template-columns: minmax(250px,.8fr) minmax(0,1.25fr); gap: clamp(34px,6vw,82px); align-items: start; }}
    .section-intro {{ margin-bottom: 34px; }}
    .section-intro p, .registry-note p, .rsvp-heading p {{ max-width: 660px; margin-top: 14px; color: var(--muted); }}
    .section-intro::after, .rsvp-heading::after {{ content: ""; display: block; width: 72px; height: 2px; margin-top: 22px; background: linear-gradient(90deg, var(--burgundy), var(--champagne), transparent); }}
    .pull-quote {{ margin-top: 28px; padding-top: 18px; border-top: 1px solid rgba(123,38,56,.18); color: var(--sage); font-weight: 900; }}
    .story-copy {{ display: grid; gap: 18px; color: rgba(43,43,43,.86); font-size: clamp(1rem,1.5vw,1.12rem); }}

    .gallery-grid {{ display: grid; grid-template-columns: 1.08fr .76fr .76fr; grid-template-rows: 220px 220px; gap: 18px; }}
    .memory-tile {{ position: relative; display: grid; align-content: space-between; min-height: 190px; padding: 20px; border-radius: 8px; overflow: hidden; box-shadow: var(--shadow); transition: transform 220ms ease, box-shadow 220ms ease; }}
    .memory-tile:hover, .event-card:hover {{ transform: translateY(-4px); box-shadow: 0 34px 86px rgba(72,41,35,.18); }}
    .memory-tile::after {{ content: ""; position: absolute; inset: 0; background: radial-gradient(circle at 20% 15%, rgba(255,250,247,.22), transparent 34%); }}
    .memory-tile span, .memory-tile strong {{ position: relative; z-index: 1; }}
    .memory-tile span {{ color: rgba(255,255,255,.68); font-family: Georgia, serif; font-size: 3.4rem; font-weight: 700; line-height: .9; }}
    .memory-tile strong {{ color: #fff; font-size: 1.04rem; line-height: 1.25; }}
    .memory-tile.feature {{ grid-row: span 2; background: linear-gradient(0deg, rgba(49,21,24,.62), rgba(49,21,24,.06)), linear-gradient(135deg,#efe2dd,#c8bfb6); }}
    .memory-tile.blush {{ background: linear-gradient(135deg,#d9b7a5,var(--burgundy)); }}
    .memory-tile.burgundy {{ background: var(--burgundy); }}
    .memory-tile.sage {{ background: linear-gradient(135deg,var(--sage),#cfc7a7); }}
    .memory-tile.ivory {{ background: linear-gradient(135deg,#f5eee8,#fffaf7); }}
    .memory-tile.ivory span, .memory-tile.ivory strong {{ color: var(--burgundy); }}

    .event-grid {{ display: grid; grid-template-columns: repeat(2,minmax(0,1fr)); gap: 18px; }}
    .event-card, div[data-testid="stVerticalBlockBorderWrapper"] {{ border: 1px solid rgba(123,38,56,.12); border-radius: 8px; background: rgba(255,250,247,.9); box-shadow: var(--shadow); }}
    .event-card {{ display: grid; gap: 22px; padding: clamp(22px,4vw,36px); transition: transform 220ms ease, box-shadow 220ms ease; }}
    .detail-list {{ display: grid; grid-template-columns: repeat(3,minmax(0,1fr)); gap: 10px; margin: 0; }}
    .detail-list div {{ min-height: 74px; padding: 12px; border-radius: 8px; background: rgba(255,255,255,.7); }}
    .detail-list dt {{ color: var(--muted); font-size: .72rem; font-weight: 900; letter-spacing: .08em; text-transform: uppercase; }}
    .detail-list dd {{ margin: 5px 0 0; font-weight: 900; }}
    .timeline {{ display: grid; gap: 12px; margin: 0; padding: 0; list-style: none; }}
    .timeline li {{ display: grid; grid-template-columns: 112px minmax(0,1fr); gap: 14px; align-items: baseline; padding-top: 12px; border-top: 1px solid rgba(123,38,56,.14); }}
    .timeline time {{ color: var(--burgundy); font-weight: 900; }}

    .registry-layout {{ display: grid; grid-template-columns: minmax(0,1fr) minmax(260px,.62fr); gap: clamp(24px,5vw,70px); align-items: center; }}
    .registry-note {{ padding: clamp(20px,4vw,34px); border-left: 4px solid var(--champagne); border-radius: 8px; background: rgba(255,250,247,.84); box-shadow: 0 18px 54px rgba(72,41,35,.08); }}
    .outline-button {{ display: inline-flex; align-items: center; min-height: 46px; margin-top: 18px; padding: 0 18px; border: 1px solid var(--burgundy); border-radius: 999px; color: var(--burgundy); font-weight: 900; }}

    div[data-testid="stVerticalBlockBorderWrapper"] {{ width: min(1120px, calc(100% - 36px)); margin: clamp(76px,9vw,124px) auto; padding: clamp(24px,5vw,46px); }}
    div[data-testid="stVerticalBlockBorderWrapper"] div[data-testid="stVerticalBlockBorderWrapper"] {{ width: 100%; margin: 18px 0 0; padding: 20px; background: rgba(248,245,242,.72); box-shadow: none; }}
    .rsvp-anchor {{ scroll-margin-top: 92px; }}
    .rsvp-heading-box {{ min-height: 228px; }}
    .rsvp-native-title h3 {{ margin: 8px 0 4px; }}
    .rsvp-native-copy {{ margin: 0 0 18px; color: var(--charcoal); }}
    div[data-testid="stTextInput"] label {{ color: var(--charcoal); font-weight: 900; }}
    div[data-testid="stTextInput"] input {{ min-height: 52px; border-radius: 999px; border: 1px solid rgba(123,38,56,.18); background: var(--ivory); color: var(--charcoal); }}
    div[data-testid="stCheckbox"] {{ padding: 9px 12px; border: 1px solid rgba(123,38,56,.16); border-radius: 8px; background: #fff; }}
    div[data-testid="stCheckbox"] label, div[data-testid="stRadio"] label, div[data-testid="stCheckbox"] label *, div[data-testid="stRadio"] label * {{ cursor: pointer !important; }}
    .site-footer {{ display: flex; align-items: center; justify-content: space-between; gap: 22px; min-height: 76px; padding: 24px clamp(18px,6vw,88px); color: #fff; background: var(--burgundy); font-weight: 900; }}
    .site-footer .couple {{ font-family: Georgia, serif; font-size: 1.35rem; white-space: nowrap; }}
    .sticky-rsvp {{ position: fixed; right: 18px; bottom: 18px; z-index: 40; min-height: 42px; padding: 0 16px; font-size: .78rem; }}
    .music-player-wrap {{
      position: fixed;
      left: 18px;
      bottom: 18px;
      z-index: 39;
      display: grid;
      gap: 6px;
      width: min(300px, calc(100vw - 36px));
      padding: 10px 12px;
      border: 1px solid rgba(214, 195, 163, 0.42);
      border-radius: 10px;
      background: rgba(255, 250, 247, 0.9);
      box-shadow: 0 18px 44px rgba(72, 41, 35, 0.12);
      backdrop-filter: blur(14px);
    }}
    .music-player-wrap span {{
      color: var(--burgundy);
      font-size: 0.68rem;
      font-weight: 900;
      letter-spacing: 0.12em;
      text-transform: uppercase;
    }}
    .music-player-wrap audio {{
      width: 100%;
      height: 34px;
      display: block;
    }}
    @keyframes weddingFloat {{ 0%, 100% {{ transform: translateY(0); }} 50% {{ transform: translateY(-5px); }} }}
    @media (max-width: 860px) {{
      .desktop-nav {{ display: none; }}
      .hero {{ min-height: 88vh; padding: 70px 18px; }}
      h1 {{ font-size: clamp(4rem,17vw,6.25rem); }}
      .section-grid, .event-grid, .registry-layout {{ grid-template-columns: 1fr; }}
      .gallery-grid {{ grid-template-columns: 1fr 1fr; grid-template-rows: 280px 170px 170px; }}
      .memory-tile.feature {{ grid-column: 1 / -1; grid-row: auto; }}
      .detail-list {{ grid-template-columns: 1fr; }}
    }}
    @media (max-width: 520px) {{
      .site-header {{ min-height: 62px; padding-inline: 14px; }}
      .brand img {{ width: 52px; }}
      .nav-button {{ min-height: 38px; padding: 0 14px; }}
      .section {{ padding-inline: 16px; }}
      .gallery-grid {{ grid-template-columns: 1fr; grid-template-rows: 260px repeat(4,150px); }}
      .timeline li {{ grid-template-columns: 1fr; }}
      .site-footer {{ display: grid; }}
      .sticky-rsvp {{ display: none; }}
      .music-player-wrap {{ left: 12px; right: 12px; bottom: 12px; width: auto; }}
    }}
    </style>
    """


def html_block(markup: str) -> None:
    st.html(markup)


def site_password() -> str:
    try:
        return str(st.secrets.get("site_password", DEFAULT_SITE_PASSWORD))
    except Exception:
        return DEFAULT_SITE_PASSWORD


def password_screen() -> bool:
    if st.session_state.get("site_unlocked"):
        return True

    logo = image_uri(ASSETS / "logo-burgundy.png")
    html_block(
        f"""
        <style>
        header[data-testid="stHeader"],
        div[data-testid="stToolbar"],
        div[data-testid="stDecoration"],
        footer {{ display: none; }}
        .block-container {{ max-width: none; padding: 0; }}
        .stApp {{
          min-height: 100vh;
          background:
            radial-gradient(circle at 24% 18%, rgba(214, 195, 163, 0.26), transparent 30%),
            linear-gradient(135deg, #fffaf7, #f8f5f2 58%, #ede6df);
          color: #2b2b2b;
          font-family: Lato, Avenir, "Segoe UI", sans-serif;
        }}
        .private-entry {{
          min-height: 56vh;
          display: grid;
          place-items: end center;
          padding: 32px 18px 18px;
        }}
        .private-card {{
          width: min(520px, 100%);
          padding: clamp(28px, 6vw, 48px);
          border: 1px solid rgba(123, 38, 56, 0.14);
          border-radius: 10px;
          background: rgba(255, 250, 247, 0.92);
          box-shadow: 0 28px 80px rgba(72, 41, 35, 0.14);
          text-align: center;
        }}
        .private-card img {{
          width: 86px;
          margin-bottom: 22px;
          filter: drop-shadow(0 8px 18px rgba(88, 23, 37, 0.12));
        }}
        .private-card .eyebrow {{
          margin: 0 0 12px;
          color: #a97724;
          font-size: 0.75rem;
          font-weight: 900;
          letter-spacing: 0.15em;
          text-transform: uppercase;
        }}
        .private-card h1 {{
          margin: 0;
          color: #7b2638;
          font-family: Georgia, "Times New Roman", serif;
          font-size: clamp(2.4rem, 8vw, 4rem);
          line-height: 0.96;
        }}
        .private-card p {{
          margin: 16px auto 0;
          max-width: 360px;
          color: #6f625c;
          line-height: 1.65;
        }}
        div[data-testid="stTextInput"] {{
          width: min(420px, calc(100% - 36px));
          margin: 0 auto;
        }}
        div[data-testid="stTextInput"] label {{
          color: #2b2b2b;
          font-weight: 900;
        }}
        div[data-testid="stTextInput"] input {{
          min-height: 52px;
          border-radius: 999px;
          border: 1px solid rgba(123, 38, 56, 0.18);
          background: #fffaf7;
          color: #2b2b2b;
        }}
        div.stButton {{
          width: min(420px, calc(100% - 36px));
          margin: 16px auto 0;
        }}
        div.stButton > button {{
          min-height: 48px;
          border: 0;
          border-radius: 999px;
          background: #7b2638;
          color: #fff;
          font-weight: 900;
          box-shadow: 0 16px 38px rgba(88, 23, 37, 0.18);
        }}
        div.stButton > button:hover {{ background: #581725; }}
        div[data-testid="stAlert"] {{
          width: min(420px, calc(100% - 36px));
          margin: 14px auto 0;
          border-radius: 8px;
        }}
        </style>
        <section class="private-entry">
          <div class="private-card">
            <img src="{logo}" alt="">
            <p class="eyebrow">Private celebration</p>
            <h1>Celine &amp; Kiran</h1>
            <p>Please enter the wedding password to view the invitation.</p>
          </div>
        </section>
        """
    )

    with st.form("wedding_password_form"):
        password = st.text_input("Password", type="password", placeholder="Enter password")
        submitted = st.form_submit_button("Enter wedding website", use_container_width=True)

    if submitted:
        if password.strip() == site_password():
            st.session_state["site_unlocked"] = True
            st.rerun()
        st.error("That password is not quite right. Please try again.")

    return False


def nav(data: dict[str, object]) -> None:
    html_block(
        f"""
        <nav class="site-header">
          <a class="brand" href="#home" aria-label="Celine and Kiran home"><img src="{data["logo_burgundy"]}" alt=""></a>
          <div class="desktop-nav">
            <a href="#story">Story</a>
            <a href="#gallery">Gallery</a>
            <a href="#details">Details</a>
            <a href="#registry">Registry</a>
          </div>
          <a class="nav-button" href="#rsvp">RSVP</a>
        </nav>
        """
    )


def hero(data: dict[str, object]) -> None:
    html_block(
        f"""
        <section class="hero" id="home">
          <div class="hero-content">
            <img class="hero-logo" src="{data["logo_ivory"]}" alt="">
            <p class="eyebrow">{escape(str(data["invitation"]))}</p>
            <h1>{escape(str(data["couple"]))}</h1>
            <p class="hero-subtitle">Are getting married</p>
            <div class="countdown"><span class="countdown-number">--</span><span class="countdown-label">Date to be announced</span></div>
          </div>
        </section>
        """
    )


def story(data: dict[str, object]) -> None:
    story_lines = data["story"]
    paragraphs = "".join(f"<p>{escape(line)}</p>" for line in story_lines)
    html_block(
        f"""
        <section class="section story" id="story">
          <div class="section-grid">
            <div>
              <p class="eyebrow">Our Love Story</p>
              <h2>Written slowly, loved deeply, timed perfectly.</h2>
              <p class="pull-quote">What we wound up building was not just a relationship, but a true partnership.</p>
            </div>
            <div class="story-copy">{paragraphs}</div>
          </div>
        </section>
        """
    )


def gallery() -> None:
    html_block(
        """
        <section class="section gallery" id="gallery">
          <div class="section-intro">
            <p class="eyebrow">Gallery</p>
            <h2>A few of our favourite memories along the way...</h2>
          </div>
          <div class="gallery-grid">
            <article class="memory-tile feature"><span>CK</span><strong>Forever starts here</strong></article>
            <article class="memory-tile blush"><span>01</span><strong>First conversations</strong></article>
            <article class="memory-tile burgundy"><span>02</span><strong>Trips that became chapters</strong></article>
            <article class="memory-tile sage"><span>03</span><strong>The proposal</strong></article>
            <article class="memory-tile ivory"><span>04</span><strong>Family, laughter, home</strong></article>
          </div>
        </section>
        """
    )


def event_card(event: dict[str, str]) -> str:
    timeline = "".join(f"<li><time>{escape(label)}</time><span>{escape(body)}</span></li>" for label, body in timeline_items(event.get("Timeline of Events", "")))
    return f"""
    <article class="event-card">
      <p class="event-label">{escape(event.get("Events", "Wedding Event"))}</p>
      <h3>{escape(event.get("Dress Code", "Elegant"))}</h3>
      <dl class="detail-list">
        <div><dt>Date</dt><dd>{escape(event.get("Date", "TBD"))}</dd></div>
        <div><dt>Venue</dt><dd>{escape(event.get("Venue", event.get("Venue ", "TBD")))}</dd></div>
        <div><dt>Location</dt><dd>{escape(event.get("Location", "TBD"))}</dd></div>
      </dl>
      <ul class="timeline">{timeline}</ul>
    </article>
    """


def details(data: dict[str, object]) -> None:
    cards = "".join(event_card(event) for event in data["events"])
    html_block(
        f"""
        <section class="section details" id="details">
          <div class="section-intro">
            <p class="eyebrow">Wedding Information</p>
            <h2>{escape(str(data["details_intro"]))}</h2>
            <p>Details will be updated here as dates and venues are confirmed.</p>
          </div>
          <div class="event-grid">{cards}</div>
        </section>
        """
    )


def registry(data: dict[str, object]) -> None:
    lines = data["registry"]
    title = lines[0] if lines else "Your presence is truly the only gift we need."
    body = lines[1] if len(lines) > 1 else "For those who have asked, a registry link will be shared soon."
    html_block(
        f"""
        <section class="section registry" id="registry">
          <div class="registry-layout">
            <div>
              <p class="eyebrow">Registry</p>
              <h2>{escape(title)}</h2>
            </div>
            <div class="registry-note">
              <p>{escape(body)}</p>
              <span class="outline-button">Registry link coming soon</span>
            </div>
          </div>
        </section>
        """
    )


def guest_controls(guest: Guest) -> None:
    seats = "1 seat" if guest.seats == 1 else f"{guest.seats} seats"
    html_block(f'<div class="rsvp-native-title"><h3>Welcome, {escape(guest.name)}</h3><p class="rsvp-native-copy">We have reserved {seats} in your honour.</p></div>')

    hindu = st.checkbox("Hindu Wedding", value=guest.hindu, disabled=not guest.hindu, key=f"hindu_{guest.name}")
    christian = st.checkbox("Christian Wedding", value=guest.christian, disabled=not guest.christian, key=f"christian_{guest.name}")
    response = st.radio("RSVP response", ["Accept", "Decline"], horizontal=True, key=f"response_{guest.name}")
    save_col, contact_col, _ = st.columns([0.24, 0.30, 0.46])
    with save_col:
        save_clicked = st.button("Save RSVP", key=f"save_{guest.name}", use_container_width=True)
    with contact_col:
        contact_clicked = st.button("Contact directly", key=f"contact_{guest.name}", use_container_width=True)
    if save_clicked or contact_clicked:
        try:
            save_rsvp(guest, response, hindu, christian, contact_clicked)
            note = " and your contact request has been noted" if contact_clicked else ""
            st.success(f"Thank you, {guest.name}. Your RSVP has been saved{note}.")
        except PermissionError:
            st.error("I could not save the RSVP because C&K.xlsx appears to be open. Please close it and try again.")


def rsvp(data: dict[str, object]) -> None:
    html_block('<div class="rsvp-anchor" id="rsvp"></div>')
    with st.container(border=True):
        left, right = st.columns([0.72, 1])
        with left:
            html_block(
                """
                <div class="rsvp-heading rsvp-heading-box">
                  <p class="eyebrow">RSVP</p>
                  <h2>Find your invitation</h2>
                  <p>Enter your full name to check the guest list and see the celebration details reserved for you.</p>
                </div>
                """
            )
        with right:
            input_col, button_col = st.columns([1, 0.28])
            with input_col:
                name = st.text_input("Full name", placeholder="e.g. Sarah Ali", key="guest_name")
            with button_col:
                st.html("<div style='height: 29px'></div>")
                lookup = st.button("Find", use_container_width=True)
            if lookup:
                st.session_state["guest_searched"] = True
                st.session_state["guest_result"] = find_guest(name, data["guests"])
            searched = st.session_state.get("guest_searched", False)
            guest = st.session_state.get("guest_result")
            if searched and guest is None:
                st.warning("We could not find that name. Please try the full name from your invitation or contact us directly.")
            elif searched and guest is not None:
                with st.container(border=True):
                    guest_controls(guest)


def footer(data: dict[str, object]) -> None:
    html_block(
        f"""
        <footer class="site-footer">
          <span>Love, laughter, and a happily ever after.</span>
          <span class="couple">{escape(str(data["couple"]))}</span>
        </footer>
        <div class="music-player-wrap">
          <span>Play our song</span>
          <audio controls preload="none" loop controlslist="nodownload">
            <source src="{MUSIC_URL}" type="audio/mpeg">
          </audio>
        </div>
        <a class="sticky-rsvp" href="#rsvp">RSVP</a>
        """
    )


def main() -> None:
    st.set_page_config(page_title="Celine & Kiran | Wedding", page_icon="CK", layout="wide", initial_sidebar_state="collapsed")
    if not password_screen():
        st.stop()

    data = load_data()
    html_block(css(data))
    nav(data)
    hero(data)
    story(data)
    gallery()
    details(data)
    registry(data)
    rsvp(data)
    footer(data)


if __name__ == "__main__":
    main()
